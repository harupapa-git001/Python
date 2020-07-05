'''
    7. openpyxl
    
    "openpyxl"はEXcelファイル（.xlsx）を扱うための外部ライブラリです。
    
    本書ではバージョン2.4.9を扱います。
    
    #インストール
    
    ■Windows
    
    pip install openpyxl
    
    ■macOS
    
    pip3 install openpyxl
    
    Excelファイルを扱えるライブラリはいくつかありますが、その中でもopenpyxlは最も人気が高いライブラリだと言っても過言ではありません。
    
    本書では基本的な操作をいくつか紹介します。

'''
#Excelファイルの作成

import openpyxl as px

#ブックの新規作成

wb = px.Workbook()

#アクティブシートの取得

ws = wb.active

#内容の更新

ws.title = "タイトル"

ws["A1"] = "これは"

ws["B2"] = "openpyxlで"

ws["C3"] = "作りました"

#ファイルの保存

wb.save("nya.xlsx")

'''
'''
#Excelファイルの読み込み

import openpyxl as px

#ブックを読み込む

wb = px.load_workbook("nya.xlsx")

#ひとつ目のシートを読み込む

ws = wb.worksheets[0]

#セルの値を取得する

print(ws["A1"].value, ws["B2"].value, ws["C3"].value)

'''

'''
#ファイルの上書きなど

import openpyxl as px

#ブックを読み込む

wb = px.load_workbook("nya.xlsx")

#名前を指定してシートを追加

ws = wb.create_sheet(title = "new_sheet")

#名前からシートを選択

ws = wb.get_sheet_by_name("new_sheet")

#シートの一覧を取得

print(wb.get_sheet_names())

#セルの結合

ws.merge_cells("A2:C3")

ws["A1"] = "結合したセル"

#ファイルの上書き保存

wb.save("nya.xlsx")

'''
    例を見れがわかると思いますが、ファイルの読み書き程度であれば補足の説明が不要なぐらい簡潔に、直感的に扱えます。
    
    なお、本書では掲載していませんが、
    
    ・枠線をつける
    ・フォントを指定する
    ・背景色をつける
    
    などの一通りの操作はもちろん可能なので、Excelを使った業務の効率化や自動化などに興味がある人はもっと詳しく調べてみると良いでしょう。
    
'''
